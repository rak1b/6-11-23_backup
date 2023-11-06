import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.http import HttpResponse

def generate_excel(headers,data):
    # Create a new workbook
    workbook = Workbook()
    # Get the active worksheet
    worksheet = workbook.active

    # Set the headers for the Excel file
    # headers = [
    #   'id','product_name','quantity','total','to_address','to_email','to_phone','to_name'
    # ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}1'] = header

    # Fetch the data from the API response
    # data = [
    #     {
    #         'id':1,
    #         "product_name":"hand Stiched",
    #         "quantity":1,
    #         "total":100,
    #         "to_address":"kathmandu",
    #         "to_phone":"9841234567",
    #         "to_email":"mail@mail.cm",
    #         "to_name":"saroj",
            
    #     },
    # {
    #         'id':2,
    #         "product_name":"hand Stiched",
    #         "quantity":1,
    #         "total":100,
    #         "to_address":"kathmandu",
    #         "to_phone":"9841234567",
    #         "to_email":"mail@mail.cm",
    #         "to_name":"saroj",
            
    #     },

    # ]

    # Write the data to the worksheet
    row_num = 2
    # for invoice in enumerate(data, 1):
    for item in data:            
        for index in range(0,len(headers)):
          print("index",index)
          worksheet.cell(row=row_num, column=index+1, value=item[headers[index]])

        row_num += 1

    # Create the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=invoice_data.xlsx'
    workbook.save(response)

    return response
