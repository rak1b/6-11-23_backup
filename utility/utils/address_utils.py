import json
import openpyxl

def load_json(path):
    with open(path) as f:
        data = json.load(f)
    return data


input_data = load_json(r'H:\Devsstream_NEW\kaaruj-backend-v2\utility\utils\address.json')

def convert_initial_json(file_name):
  output_data = {"Divisions": []}
  for division in input_data["Divisions"]:
      new_division = {
          "ID": division["ID"],
          "NAME": division["NAME"],
          "Districts": []
      }
      for district in division["Districts"]:
          new_district = {
              "ID": district["ID"],
              "NAME": district["NAME"],
              "Areas": []
          }
          for area in district["Areas"]:
              new_area = {
                  "ID": area["ID"],
                  "NAME": area["NAME"],
              }
              if area["Zone"]["NAME"] == "Inside Dhaka":
                  new_area["INSIDE_DHAKA"] = True
                  new_area["DELIVERY_CHARGE"] = 80
              else:
                  new_area["INSIDE_DHAKA"] = False
                  new_area["DELIVERY_CHARGE"] = 150
              new_district["Areas"].append(new_area)
          new_division["Districts"].append(new_district)
      output_data["Divisions"].append(new_division)

  output_file_name = f"{file_name}.json"
  with open(output_file_name, "w") as output_file:
      json.dump(output_data, output_file, indent=4)
  return output_data


def convert_formatted_data_to_excel(data,file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
 
    headers = [
        "Divisions_name", "Divisions_id", "Districts_name", "Districts_id", "Areas_name", "Areas_id", "Inside_Dhaka",
        "Delivery_Charge"
    ]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write data to the Excel sheet
    row_num = 2
    for division in data["Divisions"]:
        for district in division["Districts"]:
            for area in district["Areas"]:
                sheet.cell(row=row_num, column=1, value=division["NAME"])
                sheet.cell(row=row_num, column=2, value=division["ID"])
                sheet.cell(row=row_num, column=3, value=district["NAME"])
                sheet.cell(row=row_num, column=4, value=district["ID"])
                sheet.cell(row=row_num, column=5, value=area["NAME"])
                sheet.cell(row=row_num, column=6, value=area["ID"])
                sheet.cell(row=row_num, column=7, value=area["INSIDE_DHAKA"])
                sheet.cell(row=row_num, column=8, value=area["DELIVERY_CHARGE"])
                row_num += 1

    workbook.save(f"{file_name}.xlsx")


def convert_excel_to_json(excel_path):
    workbook = openpyxl.load_workbook("output.xlsx")
    sheet = workbook.active
    data = {
        "Divisions": []
    }
    current_division = None
    current_district = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        divisions_name, divisions_id, districts_name, districts_id, areas_name, areas_id, inside_dhaka, delivery_charge = row
        if not current_division or current_division["NAME"] != divisions_name:
            current_division = {
                "ID": divisions_id,
                "NAME": divisions_name,
                "Districts": []
            }
            data["Divisions"].append(current_division)
            current_district = None

        # Check if we need to create a new district
        if not current_district or current_district["NAME"] != districts_name:
            current_district = {
                "ID": districts_id,
                "NAME": districts_name,
                "Areas": []
            }
            current_division["Districts"].append(current_district)

        # Create an area and add it to the current district
        area = {
            "ID": areas_id,
            "NAME": areas_name,
            "INSIDE_DHAKA": bool(inside_dhaka),
            "DELIVERY_CHARGE": delivery_charge
        }
        current_district["Areas"].append(area)

    # Convert the data to JSON
    output_file_name = "reformatted_data_from_excel.json"
    with open(output_file_name, "w") as output_file:
        json.dump(data, output_file, indent=4)
    return data


convert_formatted_data_to_excel(convert_initial_json("address_data_formatted"),"address_data_formatted")
# convert_excel_to_json()